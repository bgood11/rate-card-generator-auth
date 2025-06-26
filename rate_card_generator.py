import os
from datetime import datetime
from simple_salesforce import Salesforce
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import click
from typing import Dict, List, Tuple
import json

class RateCardGenerator:
    def __init__(self, username: str, password: str, security_token: str, domain: str = 'login'):
        """Initialize Salesforce connection"""
        self.sf = Salesforce(
            username=username,
            password=password,
            security_token=security_token,
            domain=domain
        )
        
        # Remove product groupings - process each vertical separately
        self.product_groupings = None
    
    def find_retailer(self, partial_name: str, salesforce_user_id: str = None) -> List[Dict]:
        """Find retailers and retailer branches matching partial name
        Only returns accounts that have live rate cards OR branches that have their own active assigned rate cards
        
        Args:
            partial_name: Partial retailer name to search for
            salesforce_user_id: If provided, filter to only accounts owned by this user
        """
        # Build owner filter clause for reuse
        owner_filter = f" AND OwnerId = '{salesforce_user_id}'" if salesforce_user_id else ""
        
        # Use three separate queries and combine results to avoid nested semi-join restrictions
        # Query 1: Accounts that directly have live rate cards
        query1 = f"""
        SELECT Name, Id, RecordType.DeveloperName, OwnerId, Owner.Name
        FROM Account
        WHERE Name LIKE '%{partial_name}%'
            AND RecordType.DeveloperName IN ('Retailer', 'Retailer_Branch')
            AND Id IN (
                SELECT AccountId 
                FROM Opportunity 
                WHERE RecordType.DeveloperName = 'Retailer_Rate_Card' 
                AND StageName = 'Live'
            ){owner_filter}
        """
        
        # Query 2: Branch accounts that have their own active assigned rate cards
        query2 = f"""
        SELECT Name, Id, RecordType.DeveloperName, OwnerId, Owner.Name
        FROM Account
        WHERE Name LIKE '%{partial_name}%'
            AND RecordType.DeveloperName = 'Retailer_Branch'
            AND Id IN (
                SELECT Retailer__c 
                FROM Assigned_Rate_Card__c 
                WHERE Active__c = true
            ){owner_filter}
        """
        
        # Execute both queries
        results1 = self.sf.query(query1)['records']
        results2 = self.sf.query(query2)['records']
        
        # Combine and deduplicate results based on Id
        seen_ids = set()
        combined_results = []
        
        for record in results1 + results2:
            if record['Id'] not in seen_ids:
                seen_ids.add(record['Id'])
                combined_results.append(record)
        
        # Sort by name
        combined_results.sort(key=lambda x: x['Name'])
        
        return combined_results
    
    def get_rate_card_items(self, retailer_name: str) -> pd.DataFrame:
        """Get rate card items with position data using two-step approach"""
        # First check if this is a retailer branch and get parent account if needed
        account_query = f"""
        SELECT Name, RecordType.DeveloperName, Parent.Name
        FROM Account
        WHERE Name = '{retailer_name}'
        LIMIT 1
        """
        account_result = self.sf.query(account_query)
        
        # Determine which account name to use for the Opportunity query
        if account_result['records']:
            account = account_result['records'][0]
            if (account['RecordType']['DeveloperName'] == 'Retailer_Branch' and 
                account.get('Parent') and account['Parent'].get('Name')):
                # Use parent account name for branches
                opportunity_account_name = account['Parent']['Name']
                print(f"[DEBUG] Branch detected, using parent account: {opportunity_account_name}")
            else:
                # Use retailer name as-is for regular retailers
                opportunity_account_name = retailer_name
        else:
            # Fallback to retailer name if account not found
            opportunity_account_name = retailer_name
        
        # Step 1: Query Assigned_Rate_Card__c records to get positions and opportunity IDs
        arc_query = f"""
        SELECT
            Id,
            Opportunity__c,
            Prime_SubPrime__c,
            Prime_Lender_Position__c,
            Sub_Prime_Lender_Position__c,
            Opportunity__r.Lender_Company__r.Name,
            Opportunity__r.Approved_Product__r.Name,
            Opportunity__r.Shermin_Commission__c
        FROM Assigned_Rate_Card__c
        WHERE
            Retailer__r.Name = '{retailer_name}'
            AND Active__c = true
            AND Opportunity__r.Account.Name = '{opportunity_account_name}'
            AND Opportunity__r.RecordType.DeveloperName = 'Retailer_Rate_Card'
            AND Opportunity__r.StageName = 'Live'
        ORDER BY
            Opportunity__r.Lender_Company__r.Name,
            Opportunity__r.Approved_Product__r.Name
        """
        
        try:
            arc_results = self.sf.query_all(arc_query)
            print(f"[DEBUG] Found {len(arc_results['records'])} assigned rate card records")
        except Exception as e:
            print(f"[ERROR] Assigned rate card query failed: {e}")
            return self._get_rate_card_items_fallback(retailer_name, opportunity_account_name)
        
        if not arc_results['records']:
            print(f"[WARNING] No assigned rate cards found for {retailer_name}")
            return pd.DataFrame()
        
        # Step 2: For each unique opportunity, get its OpportunityLineItem records
        flattened_records = []
        processed_opportunities = set()
        
        for arc_record in arc_results['records']:
            try:
                opportunity_id = arc_record.get('Opportunity__c')
                if opportunity_id in processed_opportunities:
                    continue
                processed_opportunities.add(opportunity_id)
                
                # Get OpportunityLineItem records for this specific opportunity
                oli_query = f"""
                SELECT
                    Id,
                    OpportunityId,
                    Opportunity.Lender_Company__r.Name,
                    Opportunity.Approved_Product__r.Name,
                    Opportunity.Shermin_Commission__c,
                    Product2.Name,
                    Product2.APR__c,
                    Product2.Term__c,
                    Product2.ProductCode,
                    Product2.Deferred_Period__c,
                    Retailer_Subsidy__c,
                    Retailer_Commission__c
                FROM OpportunityLineItem
                WHERE
                    OpportunityId = '{opportunity_id}'
                    AND Active__c = true
                """
                
                oli_results = self.sf.query_all(oli_query)
                print(f"[DEBUG] Opportunity {opportunity_id}: {len(oli_results['records'])} line items")
                
                for oli_record in oli_results['records']:
                    flat_record = {
                        'Opportunity_Id': oli_record.get('OpportunityId'),
                        'Lender_Name': oli_record['Opportunity']['Lender_Company__r']['Name'] if oli_record.get('Opportunity') and oli_record['Opportunity'].get('Lender_Company__r') else None,
                        'Product_Vertical': oli_record['Opportunity']['Approved_Product__r']['Name'] if oli_record.get('Opportunity') and oli_record['Opportunity'].get('Approved_Product__r') else None,
                        'Commission': oli_record['Opportunity']['Shermin_Commission__c'] if oli_record.get('Opportunity') else None,
                        'Product_Name': oli_record['Product2']['Name'] if oli_record.get('Product2') else None,
                        'APR': oli_record['Product2']['APR__c'] if oli_record.get('Product2') else None,
                        'Term': oli_record['Product2']['Term__c'] if oli_record.get('Product2') else None,
                        'Product_Code': oli_record['Product2']['ProductCode'] if oli_record.get('Product2') else None,
                        'Deferred_Period': oli_record['Product2']['Deferred_Period__c'] if oli_record.get('Product2') else None,
                        'Subsidy': oli_record.get('Retailer_Subsidy__c', 0),
                        'Retailer_Commission': oli_record.get('Retailer_Commission__c', 0),
                        # Attach position data from the corresponding Assigned_Rate_Card__c record
                        'Prime_SubPrime': arc_record.get('Prime_SubPrime__c'),
                        'Prime_Position': arc_record.get('Prime_Lender_Position__c'),
                        'SubPrime_Position': arc_record.get('Sub_Prime_Lender_Position__c')
                    }
                    
                    # Only add records that have essential data
                    if flat_record['Lender_Name'] and flat_record['Product_Vertical']:
                        flattened_records.append(flat_record)
                    else:
                        print(f"[WARNING] Skipping record with missing lender or product vertical: {flat_record}")
                        
            except Exception as e:
                print(f"[ERROR] Failed to process ARC record: {e}")
                print(f"[ERROR] Record data: {arc_record}")
        
        print(f"[DEBUG] Total flattened records: {len(flattened_records)}")
        return pd.DataFrame(flattened_records)
    
    def _get_rate_card_items_fallback(self, retailer_name: str, opportunity_account_name: str) -> pd.DataFrame:
        """Fallback method using the original approach if main query fails"""
        print("[DEBUG] Using fallback method with separate queries")
        
        try:
            # Use the original approach: get rate items and priorities separately, then merge
            rate_items_df = self._get_rate_card_items_simple(retailer_name, opportunity_account_name)
            priorities_df = self.get_assigned_priorities(retailer_name)
            
            if rate_items_df.empty or priorities_df.empty:
                print("[WARNING] Either rate items or priorities is empty in fallback")
                return pd.DataFrame()
            
            # Merge with opportunity ID to ensure exact matching
            merged_df = pd.merge(
                rate_items_df,
                priorities_df[['Opportunity_Id', 'Lender_Name', 'Product_Vertical', 'Prime_SubPrime', 'Prime_Position', 'SubPrime_Position']],
                on=['Opportunity_Id', 'Lender_Name', 'Product_Vertical'],
                how='inner'
            )
            
            print(f"[DEBUG] Fallback merge resulted in {len(merged_df)} records")
            return merged_df
            
        except Exception as e:
            print(f"[ERROR] Fallback method failed: {e}")
            return pd.DataFrame()
    
    def _get_rate_card_items_simple(self, retailer_name: str, opportunity_account_name: str) -> pd.DataFrame:
        """Simple query to get OpportunityLineItem records"""
        query = f"""
        SELECT
            Id,
            OpportunityId,
            Opportunity.Lender_Company__r.Name,
            Opportunity.Approved_Product__r.Name,
            Opportunity.Shermin_Commission__c,
            Product2.Name,
            Product2.APR__c,
            Product2.Term__c,
            Product2.ProductCode,
            Product2.Deferred_Period__c,
            Retailer_Subsidy__c,
            Retailer_Commission__c
        FROM OpportunityLineItem
        WHERE
            Opportunity.Account.Name = '{opportunity_account_name}'
            AND Opportunity.RecordType.DeveloperName = 'Retailer_Rate_Card'
            AND Opportunity.StageName = 'Live'
            AND Active__c = true
        ORDER BY
            Opportunity.Lender_Company__r.Name,
            Opportunity.Approved_Product__r.Name
        """
        
        results = self.sf.query_all(query)
        print(f"[DEBUG] Simple query returned {len(results['records'])} opportunity line items")
        
        # Flatten nested Salesforce response
        flattened_records = []
        for record in results['records']:
            try:
                flat_record = {
                    'Opportunity_Id': record.get('OpportunityId'),
                    'Lender_Name': record['Opportunity']['Lender_Company__r']['Name'] if record.get('Opportunity') and record['Opportunity'].get('Lender_Company__r') else None,
                    'Product_Vertical': record['Opportunity']['Approved_Product__r']['Name'] if record.get('Opportunity') and record['Opportunity'].get('Approved_Product__r') else None,
                    'Commission': record['Opportunity']['Shermin_Commission__c'] if record.get('Opportunity') else None,
                    'Product_Name': record['Product2']['Name'] if record.get('Product2') else None,
                    'APR': record['Product2']['APR__c'] if record.get('Product2') else None,
                    'Term': record['Product2']['Term__c'] if record.get('Product2') else None,
                    'Product_Code': record['Product2']['ProductCode'] if record.get('Product2') else None,
                    'Deferred_Period': record['Product2']['Deferred_Period__c'] if record.get('Product2') else None,
                    'Subsidy': record.get('Retailer_Subsidy__c', 0),
                    'Retailer_Commission': record.get('Retailer_Commission__c', 0)
                }
                # Only add records that have essential data
                if flat_record['Lender_Name'] and flat_record['Product_Vertical']:
                    flattened_records.append(flat_record)
                else:
                    print(f"[WARNING] Skipping item record with missing data: {flat_record}")
            except Exception as e:
                print(f"[ERROR] Failed to process item record: {e}")
                print(f"[ERROR] Record data: {record}")
        
        return pd.DataFrame(flattened_records)
    
    def get_assigned_priorities(self, retailer_name: str) -> pd.DataFrame:
        """Query 2: Get assigned rate card priorities"""
        query = f"""
        SELECT
            Name,
            Prime_SubPrime__c,
            Prime_Lender_Position__c,
            Sub_Prime_Lender_Position__c,
            Opportunity__r.Lender_Company__r.Name,
            Opportunity__r.Approved_Product__r.Name,
            Opportunity__r.Shermin_Commission__c
        FROM Assigned_Rate_Card__c
        WHERE
            Retailer__r.Name = '{retailer_name}'
            AND Active__c = true
        ORDER BY
            Opportunity__r.Lender_Company__r.Name,
            Opportunity__r.Approved_Product__r.Name
        """
        
        results = self.sf.query_all(query)
        print(f"[DEBUG] Query returned {len(results['records'])} assigned priorities")
        
        # Debug: Log JN Bank assigned priorities
        for record in results['records']:
            lender_name = record.get('Opportunity__r', {}).get('Lender_Company__r', {}).get('Name', 'Unknown') if record.get('Opportunity__r') else 'Unknown'
            if 'JN' in lender_name.upper():
                product_vertical = record.get('Opportunity__r', {}).get('Approved_Product__r', {}).get('Name', 'Unknown') if record.get('Opportunity__r') else 'Unknown'
                prime_position = record.get('Prime_Lender_Position__c', 'None')
                subprime_position = record.get('Sub_Prime_Lender_Position__c', 'None')
                print(f"[DEBUG] Assigned: {lender_name} - {product_vertical} - Prime:{prime_position} SubPrime:{subprime_position}")
        
        # Flatten nested Salesforce response
        flattened_records = []
        for record in results['records']:
            try:
                flat_record = {
                    'Name': record.get('Name'),
                    'Opportunity_Id': record.get('Opportunity__c'),
                    'Prime_SubPrime': record.get('Prime_SubPrime__c'),
                    'Prime_Position': record.get('Prime_Lender_Position__c'),
                    'SubPrime_Position': record.get('Sub_Prime_Lender_Position__c'),
                    'Lender_Name': record['Opportunity__r']['Lender_Company__r']['Name'] if record.get('Opportunity__r') and record['Opportunity__r'].get('Lender_Company__r') else None,
                    'Product_Vertical': record['Opportunity__r']['Approved_Product__r']['Name'] if record.get('Opportunity__r') and record['Opportunity__r'].get('Approved_Product__r') else None,
                    'Commission': record['Opportunity__r']['Shermin_Commission__c'] if record.get('Opportunity__r') else None
                }
                # Only add records that have essential data
                if flat_record['Lender_Name'] and flat_record['Product_Vertical']:
                    flattened_records.append(flat_record)
                else:
                    print(f"[WARNING] Skipping priority record with missing data: {flat_record}")
            except Exception as e:
                print(f"[ERROR] Failed to process priority record: {e}")
                print(f"[ERROR] Record data: {record}")
        
        return pd.DataFrame(flattened_records)
    
    def process_rate_cards(self, retailer_name: str) -> Dict[str, pd.DataFrame]:
        """Process rate card data with embedded position information"""
        # Get data with position information already included
        rate_items_df = self.get_rate_card_items(retailer_name)
        
        # Log data counts for debugging
        print(f"\n[DEBUG] Rate items with positions found: {len(rate_items_df)}")
        
        if rate_items_df.empty:
            print(f"[WARNING] No rate card items found for {retailer_name}")
            return {}
        
        # No merge needed - position data is already included
        merged_df = rate_items_df.copy()
        
        print(f"[DEBUG] Processing data: {len(merged_df)} rows")
        print(f"[DEBUG] Unique product verticals: {merged_df['Product_Vertical'].unique()}")
        
        # Process each product vertical separately
        processed_data = {}
        for vertical in merged_df['Product_Vertical'].unique():
            if pd.isna(vertical):
                continue
                
            group_df = merged_df[merged_df['Product_Vertical'] == vertical].copy()
            
            if not group_df.empty:
                print(f"[DEBUG] Processing {vertical} with {len(group_df)} rows")
                
                # Check for missing position data
                missing_positions = group_df[
                    (group_df['Prime_Position'].isna() | (group_df['Prime_Position'] == '')) & 
                    (group_df['SubPrime_Position'].isna() | (group_df['SubPrime_Position'] == ''))
                ]
                if not missing_positions.empty:
                    print(f"[WARNING] {len(missing_positions)} rows missing position data for {vertical}")
                    # Don't skip - we want to show all rate card items
                
                # Calculate position and format data
                group_df['Position'] = group_df.apply(self._format_position, axis=1)
                group_df['Commission'] = group_df['Commission'].apply(lambda x: f"{float(x):.2f}%" if pd.notna(x) and x != 0 else "0.00%")
                
                # New subsidy logic: show negative retailer commission if subsidy is 0/blank
                def format_subsidy(row):
                    subsidy = row['Subsidy'] if pd.notna(row['Subsidy']) else 0
                    retailer_commission = row['Retailer_Commission'] if pd.notna(row['Retailer_Commission']) else 0
                    
                    if subsidy > 0:
                        return f"{float(subsidy):.2f}%"
                    elif retailer_commission > 0:
                        return f"-{float(retailer_commission):.2f}%"
                    else:
                        return "0%"
                
                group_df['Subsidy'] = group_df.apply(format_subsidy, axis=1)
                
                # Create individual rows for each unique combination
                result_data = []
                
                # Group by key characteristics to avoid duplicates while showing individual APRs
                # Don't group by Commission, Subsidy, or Retailer_Commission as these are derived fields
                grouped = group_df.groupby(['Lender_Name', 'Position', 'Term', 'Product_Code', 'Deferred_Period'])
                
                for (lender, position, term, product_code, deferred_period), term_group in grouped:
                    # Get the first row for this unique combination
                    first_row = term_group.iloc[0]
                    
                    # Get commission and subsidy from the first row
                    commission = first_row['Commission']
                    subsidy = first_row['Subsidy']
                    
                    # Format individual APR value (not a range)
                    apr_value = f"{float(first_row['APR']):.1f}%" if pd.notna(first_row['APR']) else ""
                    
                    # Format term
                    term_str = f"{int(term)} months" if pd.notna(term) else ""
                    
                    # Format deferred period
                    deferred_period_str = f"{int(deferred_period)} months" if pd.notna(deferred_period) and deferred_period > 0 else ""
                    
                    # Get product code
                    product_code_str = product_code if pd.notna(product_code) else ""
                    
                    result_data.append({
                        'Lender_Name': lender,
                        'Position': position,
                        'Shermin_Commission': commission,
                        'Term': term_str,
                        'Product_Type': product_code_str,
                        'Deferred_Period': deferred_period_str,
                        'APR_Range': apr_value,
                        'Subsidy': subsidy
                    })
                
                result_df = pd.DataFrame(result_data)
                
                # Sort by position, lender, and term
                if not result_df.empty:
                    result_df['sort_order'] = result_df['Position'].apply(self._position_sort_key)
                    result_df['term_numeric'] = result_df['Term'].str.extract(r'(\d+)').astype(float)
                    result_df = result_df.sort_values(['sort_order', 'Lender_Name', 'term_numeric']).drop(['sort_order', 'term_numeric'], axis=1)
                
                # Store with product vertical name as key
                processed_data[vertical] = result_df
                print(f"[DEBUG] Processed {vertical}: {len(result_df)} entries")
        
        return processed_data
    
    def _format_position(self, row):
        """Format position string"""
        # Handle missing position data
        if pd.isna(row.get('Prime_SubPrime')) or row.get('Prime_SubPrime') == '':
            return ''  # Return empty string for no position
            
        if row['Prime_SubPrime'] == 'Prime':
            # Handle both integer and string ordinal formats
            pos = row['Prime_Position']
            if pd.isna(pos) or pos == '':
                return ''
            if isinstance(pos, str) and pos.replace('st', '').replace('nd', '').replace('rd', '').replace('th', '').isdigit():
                return f"Prime {pos}"
            else:
                return f"Prime {self._ordinal(int(pos))}"
        else:
            # Handle both integer and string ordinal formats
            pos = row['SubPrime_Position']
            if pd.isna(pos) or pos == '':
                return ''
            if isinstance(pos, str) and pos.replace('st', '').replace('nd', '').replace('rd', '').replace('th', '').isdigit():
                return f"Sub-Prime {pos}"
            else:
                return f"Sub-Prime {self._ordinal(int(pos))}"
    
    def _ordinal(self, n):
        """Convert number to ordinal (1st, 2nd, etc.)"""
        if 10 <= n % 100 <= 20:
            suffix = 'th'
        else:
            suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')
        return f"{n}{suffix}"
    
    def _position_sort_key(self, position):
        """Create sort key for position ordering"""
        if not position or position == '':
            return (2, 999)  # Empty positions go last
            
        if 'Prime' in position and 'Sub' not in position:
            order = 0
        else:
            order = 1
        
        # Extract position number
        import re
        match = re.search(r'(\d+)', position)
        if match:
            return (order, int(match.group(1)))
        return (order, 999)
    
    
    def generate_excel(self, retailer_name: str, data: Dict[str, pd.DataFrame], output_path: str = None, hide_commissions: bool = False):
        """Generate Excel file with formatted rate cards"""
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_name = "".join(c for c in retailer_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
            output_path = f"{safe_name}_Rate_Card_{timestamp}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Rate Card Analysis"
        
        # Title
        ws['A1'] = f"{retailer_name} - Rate Card Analysis"
        ws['A1'].font = Font(bold=True, size=16)
        # Merge cells based on number of columns (8 if hiding commissions, 9 if showing)
        end_col = 'H' if hide_commissions else 'I'
        ws.merge_cells(f'A1:{end_col}1')
        
        current_row = 3
        
        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Process each product vertical group
        for group_name, df in data.items():
            if df.empty:
                continue
                
            # Group header
            ws.cell(row=current_row, column=1, value=f"{group_name} Waterfall")
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
            # Merge cells based on number of columns (8 if hiding commissions, 9 if showing)
            end_col = 'H' if hide_commissions else 'I'
            ws.merge_cells(f'A{current_row}:{end_col}{current_row}')
            current_row += 1
            
            # Table headers
            if hide_commissions:
                headers = ['Lender', 'Position', 'Term', 'Product Type', 'Deferred Period', 'APR Range', 'Subsidy', 'Changes']
            else:
                headers = ['Lender', 'Position', 'Shermin Commission', 'Term', 'Product Type', 'Deferred Period', 'APR Range', 'Subsidy', 'Changes']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            current_row += 1
            
            # Data rows
            for _, row in df.iterrows():
                if hide_commissions:
                    values = [
                        row.get('Lender_Name', ''),
                        row.get('Position', ''),
                        row.get('Term', ''),
                        row.get('Product_Type', ''),
                        row.get('Deferred_Period', ''),
                        row.get('APR_Range', ''),
                        row.get('Subsidy', ''),
                        ''  # Changes column - empty for user input
                    ]
                else:
                    values = [
                        row.get('Lender_Name', ''),
                        row.get('Position', ''),
                        row.get('Shermin_Commission', ''),
                        row.get('Term', ''),
                        row.get('Product_Type', ''),
                        row.get('Deferred_Period', ''),
                        row.get('APR_Range', ''),
                        row.get('Subsidy', ''),
                        ''  # Changes column - empty for user input
                    ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.border = border
                    # Center align commission and subsidy columns based on hide_commissions setting
                    if hide_commissions:
                        if col in [7]:  # Subsidy column when commission is hidden
                            cell.alignment = Alignment(horizontal='center')
                    else:
                        if col in [3, 8]:  # Commission and Subsidy columns when commission is shown
                            cell.alignment = Alignment(horizontal='center')
                    
                    # Add dropdown for Changes column (last column)
                    if col == len(values):  # Changes column
                        from openpyxl.worksheet.datavalidation import DataValidation
                        dv = DataValidation(type="list", formula1='"Disable,New"', allow_blank=True)
                        dv.add(cell)
                        ws.add_data_validation(dv)
            
                current_row += 1
            
            # Add spacing between tables
            current_row += 2
        
        # Adjust column widths based on hide_commissions setting
        if hide_commissions:
            # Redistribute commission column width to other columns when hidden
            column_widths = [28, 18, 15, 18, 18, 15, 12, 15]  # Without commission column
        else:
            column_widths = [25, 15, 12, 12, 15, 15, 12, 10, 12]  # With commission column
        
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Save file
        wb.save(output_path)
        return output_path

@click.command()
@click.option('--retailer', '-r', required=True, help='Retailer name (partial match supported)')
@click.option('--username', '-u', envvar='SF_USERNAME', help='Salesforce username')
@click.option('--password', '-p', envvar='SF_PASSWORD', help='Salesforce password')
@click.option('--token', '-t', envvar='SF_TOKEN', help='Salesforce security token')
@click.option('--domain', '-d', default='login', help='Salesforce domain (login/test)')
@click.option('--output', '-o', help='Output file path')
def generate_rate_card(retailer, username, password, token, domain, output):
    """Generate rate card analysis for a retailer"""
    
    # Initialize generator
    try:
        generator = RateCardGenerator(username, password, token, domain)
    except Exception as e:
        click.echo(f"Error connecting to Salesforce: {e}", err=True)
        return
    
    # Find retailer
    click.echo(f"Searching for retailers matching '{retailer}'...")
    retailers = generator.find_retailer(retailer)
    
    if not retailers:
        click.echo(f"No retailers found matching '{retailer}'", err=True)
        return
    
    if len(retailers) > 1:
        click.echo("Multiple retailers found:")
        for i, r in enumerate(retailers[:10]):  # Show max 10
            click.echo(f"  {i+1}. {r['Name']}")
        
        choice = click.prompt("Select retailer number", type=int)
        if choice < 1 or choice > len(retailers):
            click.echo("Invalid selection", err=True)
            return
        
        selected_retailer = retailers[choice-1]['Name']
    else:
        selected_retailer = retailers[0]['Name']
    
    click.echo(f"\nGenerating rate card for: {selected_retailer}")
    
    try:
        # Process rate cards
        click.echo("Fetching rate card data...")
        rate_card_data = generator.process_rate_cards(selected_retailer)
        
        # Generate Excel
        click.echo("Generating Excel file...")
        output_file = generator.generate_excel(selected_retailer, rate_card_data, output)
        
        click.echo(f"\n✅ Rate card generated successfully: {output_file}")
        
    except Exception as e:
        click.echo(f"Error generating rate card: {e}", err=True)
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    generate_rate_card()
